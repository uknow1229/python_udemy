# Excelファイルの扱い方

import openpyxl # エクセルを扱えるライブラリ

wb=openpyxl.load_workbook("出社在宅集計表_人事部.xlsx")

ws=wb["4月"]

value = ws.cell(row=2, column=2).value

ws.cell(row=5, column=5).value = 1000

wb.save("test.xlsx")

print(value)